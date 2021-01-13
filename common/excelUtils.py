import xlrd
from openpyxl import load_workbook
from common.logger import Mylog
from common.path_os import file_path

# 设置日志
logger = Mylog('Doexcel').getlog()


class ExcelUtils:
    """Excel文件操作"""
    casefile = file_path().get_case_path()

    def __init__(self, sheet_name):
        try:
            self.wb = xlrd.open_workbook(self.casefile)
            self.ws = self.wb.sheet_by_name(sheet_name)
        except FileNotFoundError as e:
            logger.error('未找到文件:' + self.casefile + '！')

    def _getcellvalues(self, row, col):
        # 读取sheet中行列的值
        mergerd_info = self.ws.merged_cells
        value = self.ws.cell_value(row, col)
        # 处理合并单元格
        for (rlow, rhigh, clow, chigh) in mergerd_info:
            if rlow <= row < rhigh:
                if clow <= col < chigh:
                    value = self.ws.cell_value(rlow, clow)
                    break
                else:
                    value = self.ws.cell_value(row, col)
            else:
                value = self.ws.cell_value(row, col)
        return value

    def ReadExcel(self):
        # 读取excel，输出dict：{[case1],[case2],[case3]}
        test_data = []
        for row in range(1, self.ws.nrows):
            row_data = {}
            for col in range(0, self.ws.ncols):
                TheKey = str(self.ws.cell_value(0, col))
                row_data[TheKey] = self._getcellvalues(row, col)
            test_data.append(row_data)
        logger.info('Excel读取测试用例：{}'.format(test_data))
        return test_data

    @staticmethod
    def WriteExcel(sheet_name, row, value='None', col=11):
        # 打开excel文件
        wb = load_workbook(ExcelUtils.casefile)
        # 打开excel表单
        sheet = wb[sheet_name]
        # 对应行列写入数据
        sheet.cell(row, col).value = value
        try:
            wb.save(ExcelUtils.casefile)
            logger.info('数据成功！:{} -> {}[{}] '.format(value, sheet_name, row))
        except Exception as e:
            logger.error('数据失败！:{} -> {}[{}],失败原因：{} '.format(value, sheet_name, row, e))

    @staticmethod
    def get_excel_base_wirte_conf():
        '''读取excel中Base数据，并写入conf文件中'''
        wb = load_workbook(file_path().get_case_path())
        sheet = wb['Base']
        CaseBase = {}
        for i in range(1, sheet.max_row + 1):
            CaseBase[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
        logger.info('读取Base数据:{}'.format(CaseBase))
        return CaseBase


if __name__ == '__main__':
    from common.configUtils import ConfigUtils

    sheetname = ConfigUtils().get_config_value('excel_case', 'SheetName')
    caselist = ExcelUtils(sheetname).ReadExcel()
    print(caselist)
    # get_excel_base_wite_conf()
