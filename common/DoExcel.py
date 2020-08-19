from openpyxl import load_workbook
import json
from common.DoConfig import *
from common.logger import Mylog
from common.path_os import file_path

logger = Mylog()


class doexcel:
    host = ReadConfig.read_config('Base', 'host')
    casefile = file_path().get_case_path()

    @classmethod
    def read_excel(cls, sheetname, case):
        wb = load_workbook(cls.casefile)
        test_case = []
        sheet = wb[sheetname]
        if case == 'all':
            for i in range(2, sheet.max_row + 1):
                # y为执行，n为不执行
                if sheet.cell(i, 2).value and sheet.cell(i, 2).value.lower() == 'y':
                    row_data = {}
                    row_data['sheetname'] = sheetname  # sheet名称
                    row_data['caseid'] = sheet.cell(i, 1).value  # 用例ID
                    row_data['tag'] = sheet.cell(i, 3).value  # 用例名称
                    row_data['url'] = cls.host + str(sheet.cell(i, 4).value)  # host存储在conf文件中，接口url
                    row_data['method'] = sheet.cell(i, 5).value  # 请求方式
                    row_data['header'] = sheet.cell(i, 6).value  # 请求头
                    row_data['params'] = sheet.cell(i, 7).value  # 入参参数
                    row_data['response'] = sheet.cell(i, 8).value  # 返参提取字段
                    row_data['sql'] = sheet.cell(i, 9).value  # DB 数据验证sql
                    row_data['expected'] = sheet.cell(i, 10).value  # 预期结果
                    # 请求头添加模拟客户端
                    if row_data['header']:
                        row_data['header']=eval(row_data['header'])
                        row_data['header'][
                            'User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'
                    test_case.append(row_data)
        elif isinstance(case, list):
            for row in case:
                row += 1  # 行数+1 == caseId
                if sheet.cell(row, 2).value and sheet.cell(row, 2).value.lower() == 'y':
                    row_data = {}
                    row_data['sheetname'] = sheetname  # sheet名称
                    row_data['caseid'] = sheet.cell(row, 1).value  # 用例ID
                    row_data['tag'] = sheet.cell(row, 3).value  # 用例名称
                    row_data['url'] = cls.host + str(sheet.cell(row, 4).value)  # host存储在conf文件中，接口url
                    row_data['method'] = sheet.cell(row, 5).value  # 请求方式
                    row_data['header'] = sheet.cell(row, 6).value  # 请求头
                    row_data['params'] = sheet.cell(row, 7).value  # 入参参数
                    row_data['response'] = sheet.cell(row, 8).value  # 返参提取字段
                    row_data['sql'] = sheet.cell(row, 9).value  # DB 数据验证sql
                    row_data['expected'] = sheet.cell(row, 10).value  # 预期结果
                    # 请求头添加模拟客户端
                    if row_data['header']:
                        logger.info('11111:{},type:{}'.format(row_data['header'], type(eval(row_data['header']))))
                        row_data['header'] = eval(row_data['header'])
                        row_data['header'][
                            'User-Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'
                    test_case.append(row_data)
        else:
            logger.error('请检查conf -> sheetname 格式！')
        return test_case

    @classmethod
    def write_excel(cls, sheetname, row, value='None', col=11):
        # 读取文件
        wb = load_workbook(cls.casefile)
        # 打开excel表单
        sheet = wb[sheetname]
        # 对应行列写入数据
        sheet.cell(row, col).value = value
        # 保存文件
        try:
            wb.save(cls.casefile)
            logger.info('用例结果成功写入：{}'.format(value))
        except Exception as e:
            logger.error('用例执行结果写入失败：{}'.format(e))


import re


def get_case_data(sheetname):

    '''解析excel_case,获取casedata,替换${},转换json格式'''
    case_data = []
    # 读取ini得到{excel表格名称 ：[caseId]}
    if isinstance(sheetname, dict):
        for k, v in sheetname.items():
            if ('[' and ']') in v:
                # 获取部分case
                data = doexcel.read_excel(k, eval(v))
            elif 'off' == v:
                continue
            else:
                # 获取全部case
                data = doexcel.read_excel(k, v)
            case_data += data
        return case_data
    else:
        logger.error('请检查配置文件 -> excel_case格式！')


def load_params(params):
    if ('{' or '}') in params:
        params = eval(params)
        # print('++++++++++++++++++++++', params, type(params))
        return params
    else:
        return params


def get_excel_base_wite_conf():
    '''读取excel中Base数据，并写入conf文件中'''
    wb = load_workbook(file_path().get_case_path())
    sheet = wb['Base']
    CaseBase = {}
    for i in range(1, sheet.max_row + 1):
        CaseBase[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
        WriteConfig.write_config('Base', sheet.cell(i, 1).value, sheet.cell(i, 2).value)
    logger.info('读取Base数据:{}'.format(CaseBase))
    return CaseBase


if __name__ == '__main__':
    from common.DoConfig import ReadConfig

    sheetname = ReadConfig.read_config_options_dict('excel_case')
    caselist = get_case_data(sheetname)
    print(caselist)
    # get_excel_base_wite_conf()
